import openpyxl as excel
import string                                               # String used for alphanumeric values


wb_name = input('Workbook file name here: ') + '.xlsx'      # Select and load Excel Workbook
wb = excel.load_workbook(wb_name)

sheet_name = input('Sheet name here (Case Sensitive): ')    # Select and load Excel Sheet in Workbook
sheet = wb[sheet_name]

cell_name = input('Starting cell name here: ').lower()      # Select and load Excel Cell in Sheet
cell_column = string.ascii_letters.index(cell_name[0]) + 1  # Must convert into nums before openpyxl runs
cell_row = int(cell_name[1])                                # Openpyxl start: 1, ascii_letters start: 0

corrected_value = input('Corrected value here(decimal): ')
save_name = input('Save Name Here: ') 

for row in range(cell_row, sheet.max_row + 1):
    cell = sheet.cell(row, cell_column)
    corrected_price = cell.value * float(corrected_value)
    corrected_price_cell = sheet.cell(row, cell_column + 1)
    corrected_price_cell.value = corrected_price


wb.save(save_name +'.xlsx')

